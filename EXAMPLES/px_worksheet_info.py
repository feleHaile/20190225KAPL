#!/usr/bin/env python
# (c)2015 John Strickler

import openpyxl as px

def main():
    """program entry point"""
    wb = px.load_workbook('../DATA/presidents.xlsx')
    ws = wb['US Presidents']

    print(ws.dimensions)
    print(ws.min_column)
    print(ws.min_row)
    print(ws.max_column)
    print(ws.max_row)
    # returns value for specified cell
    print(ws.cell(row=2, column=3).value, ws.cell(row=2, column=2).value, '\n')
    # ws.rows is all rows
    # ws.columns is all columns
    for row in ws.rows:
        for col in row[2:0:-1]:
            print(col.value, end=' ')
        print()

if __name__ == '__main__':
    main()
